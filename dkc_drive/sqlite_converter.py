"""
スプレッドシートSQLite中間形式コンバーター

Excel/CSVファイルをSQLiteに変換し、高速な読み書きを実現する。
元ファイル（.xlsx/.csv）は保持し、.sqlite を作業コピーとして使用。
"""
import os
import re
import csv
import json
import sqlite3
import hashlib
import logging
import io
import base64
from contextlib import contextmanager

from openpyxl import load_workbook, Workbook
from openpyxl.styles import PatternFill, Border, Side, Font, Alignment, Color
from openpyxl.utils import get_column_letter
from openpyxl.drawing.image import Image as OpenpyxlImage
from openpyxl.utils.cell import column_index_from_string
from openpyxl.chart import BarChart, LineChart, PieChart, DoughnutChart, RadarChart, Reference

logger = logging.getLogger(__name__)


SCHEMA_SQL = """
CREATE TABLE IF NOT EXISTS metadata (key TEXT PRIMARY KEY, value TEXT);
CREATE TABLE IF NOT EXISTS sheets (
    sheet_name TEXT PRIMARY KEY,
    sheet_order INTEGER,
    max_row INTEGER DEFAULT 1,
    max_col INTEGER DEFAULT 1,
    auto_filter TEXT,
    col_widths TEXT DEFAULT '{}',
    row_heights TEXT DEFAULT '{}'
);
CREATE TABLE IF NOT EXISTS cells (
    sheet_name TEXT,
    row INTEGER,
    col INTEGER,
    value TEXT,
    style TEXT,
    PRIMARY KEY (sheet_name, row, col)
);
CREATE TABLE IF NOT EXISTS merges (
    id INTEGER PRIMARY KEY AUTOINCREMENT,
    sheet_name TEXT,
    start_row INTEGER,
    start_col INTEGER,
    end_row INTEGER,
    end_col INTEGER
);
CREATE TABLE IF NOT EXISTS images (
    id INTEGER PRIMARY KEY AUTOINCREMENT,
    sheet_name TEXT,
    image_index INTEGER,
    row INTEGER,
    col INTEGER,
    offset_x INTEGER DEFAULT 0,
    offset_y INTEGER DEFAULT 0,
    width INTEGER DEFAULT 200,
    height INTEGER DEFAULT 200,
    content_type TEXT DEFAULT 'image/png',
    image_data BLOB,
    image_type TEXT DEFAULT NULL,
    stamped_at TEXT DEFAULT NULL,
    stamped_by TEXT DEFAULT NULL,
    rotation INTEGER DEFAULT 0
);
CREATE TABLE IF NOT EXISTS shapes (
    id INTEGER PRIMARY KEY AUTOINCREMENT,
    sheet_name TEXT,
    shape_type TEXT NOT NULL,
    row INTEGER,
    col INTEGER,
    offset_x INTEGER DEFAULT 0,
    offset_y INTEGER DEFAULT 0,
    width INTEGER DEFAULT 120,
    height INTEGER DEFAULT 80,
    fill_color TEXT DEFAULT '#4285f4',
    stroke_color TEXT DEFAULT '#333333',
    stroke_width INTEGER DEFAULT 2,
    opacity REAL DEFAULT 1.0,
    rotation INTEGER DEFAULT 0
);
CREATE TABLE IF NOT EXISTS charts (
    id INTEGER PRIMARY KEY AUTOINCREMENT,
    sheet_name TEXT,
    chart_type TEXT NOT NULL,
    min_row INTEGER,
    max_row INTEGER,
    min_col INTEGER,
    max_col INTEGER,
    row INTEGER,
    col INTEGER,
    offset_x INTEGER DEFAULT 0,
    offset_y INTEGER DEFAULT 0,
    width INTEGER DEFAULT 480,
    height INTEGER DEFAULT 320,
    title TEXT DEFAULT '',
    options TEXT DEFAULT '{}'
);
"""


def _normalize_rgb(rgb, skip_black=False):
    """'AARRGGBB' or 'RRGGBB' → '#RRGGBB' or None"""
    if not isinstance(rgb, str) or len(rgb) < 6:
        return None
    if len(rgb) == 8:
        rgb = rgb[2:]
    if skip_black and rgb == '000000':
        return None
    return f'#{rgb}'


def _parse_image_anchor(anchor):
    """画像アンカーから位置情報を解析"""
    if hasattr(anchor, '_from'):
        row = anchor._from.row
        col = anchor._from.col
        offset_x = int(anchor._from.colOff / 9525) if hasattr(anchor._from, 'colOff') else 0
        offset_y = int(anchor._from.rowOff / 9525) if hasattr(anchor._from, 'rowOff') else 0
        return row, col, offset_x, offset_y
    elif isinstance(anchor, str):
        from openpyxl.utils.cell import coordinate_from_string
        col_letter, row_num = coordinate_from_string(anchor)
        return int(row_num) - 1, column_index_from_string(col_letter) - 1, 0, 0
    return None


def _calc_twocell_image_size(anchor, ws, col_widths_px):
    """TwoCellAnchorからセル範囲ベースで実際のピクセルサイズを計算"""
    if not hasattr(anchor, '_from') or not hasattr(anchor, 'to'):
        return None, None
    f = anchor._from
    t = anchor.to
    default_row_h = ws.sheet_format.defaultRowHeight or 15
    default_row_px = default_row_h * 4 / 3  # pt -> px

    # 幅: from.col ~ to.col の列幅を合算
    width = 0
    for c in range(f.col, t.col):
        width += col_widths_px.get(c, col_widths_px.get('default', 80))
    width = width - int(f.colOff / 9525) + int(t.colOff / 9525)

    # 高さ: from.row ~ to.row の行高を合算
    height = 0
    for r in range(f.row, t.row):
        rd = ws.row_dimensions.get(r + 1)
        if rd and rd.height:
            height += rd.height * 4 / 3  # pt -> px
        else:
            height += default_row_px
    height = height - int(f.rowOff / 9525) + int(t.rowOff / 9525)

    return max(1, round(width)), max(1, round(height))


def _get_image_content_type(image):
    """画像オブジェクトからMIMEタイプを返す"""
    if hasattr(image, 'format') and image.format:
        return f'image/{image.format.lower()}'
    if hasattr(image, 'path') and image.path:
        ext = os.path.splitext(image.path)[1].lower()
        mime_map = {'.jpg': 'image/jpeg', '.jpeg': 'image/jpeg',
                    '.gif': 'image/gif', '.webp': 'image/webp'}
        return mime_map.get(ext, 'image/png')
    return 'image/png'


class SpreadsheetSQLite:
    """SQLite中間形式によるスプレッドシート操作クラス"""

    # シート名に紐づく全テーブル（cascade操作用）
    _SHEET_TABLES = ('sheets', 'cells', 'merges', 'images', 'shapes', 'charts')

    def __init__(self, filepath):
        """
        Args:
            filepath: 元ファイル（.xlsx/.csv）のパス
        """
        self.filepath = filepath
        self.sqlite_path = filepath + '.sqlite'
        self.ext = os.path.splitext(filepath)[1].lower()

    @contextmanager
    def _connect(self):
        """SQLite接続のコンテキストマネージャ"""
        conn = sqlite3.connect(self.sqlite_path)
        conn.execute("PRAGMA journal_mode=WAL")
        conn.row_factory = sqlite3.Row
        try:
            self._migrate_images_table(conn)
            self._migrate_shapes_table(conn)
            self._migrate_charts_table(conn)
            yield conn
            conn.commit()
        except Exception:
            conn.rollback()
            raise
        finally:
            conn.close()

    def _init_db(self, conn):
        """スキーマを初期化"""
        conn.executescript(SCHEMA_SQL)

    def _migrate_images_table(self, conn):
        """既存DBにimage_type/stamped_at/stamped_byカラムを追加（なければ）"""
        try:
            rows = conn.execute("PRAGMA table_info(images)").fetchall()
        except Exception:
            return
        if not rows:
            return
        cols = {row[1] for row in rows}
        for col, typ, default in [
            ('image_type', 'TEXT', 'NULL'),
            ('stamped_at', 'TEXT', 'NULL'),
            ('stamped_by', 'TEXT', 'NULL'),
            ('rotation', 'INTEGER', '0'),
        ]:
            if col not in cols:
                conn.execute(f"ALTER TABLE images ADD COLUMN {col} {typ} DEFAULT {default}")

    def _migrate_shapes_table(self, conn):
        """既存DBにshapesテーブルを追加（なければ）"""
        try:
            conn.execute("SELECT 1 FROM shapes LIMIT 1")
        except sqlite3.OperationalError:
            conn.execute("""
                CREATE TABLE IF NOT EXISTS shapes (
                    id INTEGER PRIMARY KEY AUTOINCREMENT,
                    sheet_name TEXT,
                    shape_type TEXT NOT NULL,
                    row INTEGER,
                    col INTEGER,
                    offset_x INTEGER DEFAULT 0,
                    offset_y INTEGER DEFAULT 0,
                    width INTEGER DEFAULT 120,
                    height INTEGER DEFAULT 80,
                    fill_color TEXT DEFAULT '#4285f4',
                    stroke_color TEXT DEFAULT '#333333',
                    stroke_width INTEGER DEFAULT 2,
                    opacity REAL DEFAULT 1.0,
                    rotation INTEGER DEFAULT 0
                )
            """)

    def _migrate_charts_table(self, conn):
        """既存DBにchartsテーブルを追加（なければ）"""
        try:
            conn.execute("SELECT 1 FROM charts LIMIT 1")
        except sqlite3.OperationalError:
            conn.execute("""
                CREATE TABLE IF NOT EXISTS charts (
                    id INTEGER PRIMARY KEY AUTOINCREMENT,
                    sheet_name TEXT,
                    chart_type TEXT NOT NULL,
                    min_row INTEGER,
                    max_row INTEGER,
                    min_col INTEGER,
                    max_col INTEGER,
                    row INTEGER,
                    col INTEGER,
                    offset_x INTEGER DEFAULT 0,
                    offset_y INTEGER DEFAULT 0,
                    width INTEGER DEFAULT 480,
                    height INTEGER DEFAULT 320,
                    title TEXT DEFAULT '',
                    options TEXT DEFAULT '{}'
                )
            """)

    # ============================================================
    # 変換: 元ファイル → SQLite
    # ============================================================

    def ensure_sqlite(self):
        """SQLiteファイルがなければ元ファイルから変換し、元ファイルを削除する。

        Returns:
            bool: 変換が行われたかどうか
        """
        if os.path.exists(self.sqlite_path):
            # 元ファイルが残っていれば削除（ダウンロード後の残留ファイル）
            if os.path.exists(self.filepath):
                os.remove(self.filepath)
            return False

        if not os.path.exists(self.filepath):
            raise FileNotFoundError(f'ファイルが見つかりません: {self.filepath}')

        if self.ext == '.csv':
            self.csv_to_sqlite()
        else:
            self.xlsx_to_sqlite()

        # 元ファイルを削除（以降はSQLiteのみ使用）
        os.remove(self.filepath)
        return True

    def xlsx_to_sqlite(self):
        """Excelファイル → SQLite変換"""
        wb = load_workbook(self.filepath)
        try:
            with self._connect() as conn:
                self._init_db(conn)

                conn.execute("INSERT OR REPLACE INTO metadata VALUES (?, ?)",
                             ('original_filename', os.path.basename(self.filepath)))
                conn.execute("INSERT OR REPLACE INTO metadata VALUES (?, ?)",
                             ('original_format', 'xlsx'))
                conn.execute("INSERT OR REPLACE INTO metadata VALUES (?, ?)",
                             ('dirty', '0'))

                for order, sheet_name in enumerate(wb.sheetnames):
                    ws = wb[sheet_name]
                    self._import_sheet(conn, ws, sheet_name, order)
        finally:
            wb.close()

    def _import_sheet(self, conn, ws, sheet_name, order):
        """openpyxlワークシートをSQLiteにインポート"""
        max_row = ws.max_row or 1
        max_col = ws.max_column or 1

        # 列幅
        col_widths = {}
        default_col_width = ws.sheet_format.defaultColWidth
        if default_col_width is not None:
            col_widths['default'] = round(default_col_width * 7)
        for col_letter, col_dim in ws.column_dimensions.items():
            if col_dim.width:
                col_widths[column_index_from_string(col_letter) - 1] = round(col_dim.width * 7)

        # 行高
        row_heights = {}
        default_row_height = ws.sheet_format.defaultRowHeight
        if default_row_height is not None:
            row_heights['default'] = round(default_row_height)
        for row_idx, row_dim in ws.row_dimensions.items():
            if row_dim.height:
                row_heights[row_idx - 1] = round(row_dim.height)

        # オートフィルター
        auto_filter = None
        if ws.auto_filter and ws.auto_filter.ref:
            auto_filter = json.dumps({'ref': ws.auto_filter.ref})

        conn.execute(
            "INSERT INTO sheets VALUES (?, ?, ?, ?, ?, ?, ?)",
            (sheet_name, order, max_row, max_col,
             auto_filter, json.dumps(col_widths), json.dumps(row_heights))
        )

        # セル
        cell_batch = []
        for row in ws.iter_rows(min_row=1, max_row=max_row, max_col=max_col):
            for cell in row:
                style = self._extract_cell_style(cell)
                if cell.value is None and not style:
                    continue
                r = cell.row - 1
                c = cell.column - 1
                val = '' if cell.value is None else str(cell.value)
                cell_batch.append((sheet_name, r, c, val, json.dumps(style) if style else None))

        if cell_batch:
            conn.executemany(
                "INSERT INTO cells (sheet_name, row, col, value, style) VALUES (?, ?, ?, ?, ?)",
                cell_batch
            )

        # セル結合
        for m in ws.merged_cells.ranges:
            conn.execute(
                "INSERT INTO merges (sheet_name, start_row, start_col, end_row, end_col) VALUES (?, ?, ?, ?, ?)",
                (sheet_name, m.min_row - 1, m.min_col - 1, m.max_row - 1, m.max_col - 1)
            )

        # 画像
        for idx, image in enumerate(ws._images):
            try:
                pos = _parse_image_anchor(image.anchor)
                if pos is None:
                    continue
                img_row, img_col, off_x, off_y = pos
                content_type = _get_image_content_type(image)
                image_data = image._data()

                # TwoCellAnchorの場合、セル範囲から実際のサイズを計算
                tw, th = _calc_twocell_image_size(image.anchor, ws, col_widths)
                img_w = tw if tw else (int(image.width) if image.width else 200)
                img_h = th if th else (int(image.height) if image.height else 200)

                conn.execute(
                    "INSERT INTO images (sheet_name, image_index, row, col, offset_x, offset_y, width, height, content_type, image_data) "
                    "VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?)",
                    (sheet_name, idx, img_row, img_col, off_x, off_y,
                     img_w, img_h,
                     content_type, image_data)
                )
            except Exception as e:
                logger.error("画像インポートエラー: %s", e)

    def _extract_cell_style(self, cell):
        """openpyxlセルからスタイル辞書を抽出"""
        style = {}

        # 背景色
        if cell.fill and cell.fill.fgColor and cell.fill.fgColor.rgb:
            color = _normalize_rgb(cell.fill.fgColor.rgb, skip_black=True)
            if color:
                style['backgroundColor'] = color

        # 罫線
        border = {}
        if cell.border:
            for side_name in ['left', 'right', 'top', 'bottom']:
                side = getattr(cell.border, side_name)
                if side and side.style:
                    border_color = '#000000'
                    if side.color and side.color.rgb:
                        c = _normalize_rgb(side.color.rgb)
                        if c:
                            border_color = c
                    border[side_name] = {'style': side.style, 'color': border_color}
        if border:
            style['border'] = border

        # フォント
        if cell.font:
            if cell.font.bold:
                style['fontWeight'] = 'bold'
            if cell.font.italic:
                style['fontStyle'] = 'italic'
            if cell.font.color and cell.font.color.rgb:
                color = _normalize_rgb(cell.font.color.rgb)
                if color:
                    style['color'] = color
            if cell.font.size:
                style['fontSize'] = cell.font.size

        # 配置
        if cell.alignment:
            if cell.alignment.horizontal:
                style['textAlign'] = cell.alignment.horizontal
            if cell.alignment.vertical:
                style['verticalAlign'] = cell.alignment.vertical

        return style

    def csv_to_sqlite(self):
        """CSVファイル → SQLite変換"""
        rows = []
        encoding = 'utf-8'
        for enc in ['utf-8', 'utf-8-sig', 'shift_jis', 'cp932']:
            try:
                with open(self.filepath, 'r', encoding=enc) as f:
                    reader = csv.reader(f)
                    rows = list(reader)
                encoding = enc
                break
            except (UnicodeDecodeError, UnicodeError):
                continue

        with self._connect() as conn:
            self._init_db(conn)

            conn.execute("INSERT OR REPLACE INTO metadata VALUES (?, ?)",
                         ('original_filename', os.path.basename(self.filepath)))
            conn.execute("INSERT OR REPLACE INTO metadata VALUES (?, ?)",
                         ('original_format', 'csv'))
            conn.execute("INSERT OR REPLACE INTO metadata VALUES (?, ?)",
                         ('dirty', '0'))
            conn.execute("INSERT OR REPLACE INTO metadata VALUES (?, ?)",
                         ('csv_encoding', encoding))

            sheet_name = 'Sheet1'
            max_row = len(rows) if rows else 1
            max_col = max((len(r) for r in rows), default=1) if rows else 1

            conn.execute(
                "INSERT INTO sheets VALUES (?, ?, ?, ?, ?, ?, ?)",
                (sheet_name, 0, max_row, max_col, None, '{}', '{}')
            )

            cell_batch = []
            for r, row in enumerate(rows):
                for c, val in enumerate(row):
                    if val:
                        cell_batch.append((sheet_name, r, c, val, None))

            if cell_batch:
                conn.executemany(
                    "INSERT INTO cells (sheet_name, row, col, value, style) VALUES (?, ?, ?, ?, ?)",
                    cell_batch
                )

    # ============================================================
    # 変換: SQLite → 元ファイル（ダウンロード用）
    # ============================================================

    def sqlite_to_xlsx(self):
        """SQLiteからExcelファイルを復元"""
        wb = Workbook()
        if wb.sheetnames:
            del wb[wb.sheetnames[0]]

        with self._connect() as conn:
            sheets = conn.execute(
                "SELECT * FROM sheets ORDER BY sheet_order"
            ).fetchall()

            for sheet in sheets:
                ws = wb.create_sheet(title=sheet['sheet_name'])
                sn = sheet['sheet_name']

                self._export_col_widths(ws, sheet)
                self._export_row_heights(ws, sheet)

                if sheet['auto_filter']:
                    af = json.loads(sheet['auto_filter'])
                    ws.auto_filter.ref = af.get('ref')

                self._export_cells(conn, ws, sn)
                self._export_merges(conn, ws, sn)
                self._export_shapes(conn, ws, sn)
                self._export_images(conn, ws, sn)
                self._export_charts(conn, ws, sn)

        wb.save(self.filepath)
        wb.close()

    @staticmethod
    def _export_col_widths(ws, sheet):
        """列幅をワークシートに反映"""
        col_widths = json.loads(sheet['col_widths'] or '{}')
        default_col_px = col_widths.pop('default', None)
        if default_col_px is not None:
            ws.sheet_format.defaultColWidth = default_col_px / 7
        for col_str, px_width in col_widths.items():
            col_letter = get_column_letter(int(col_str) + 1)
            ws.column_dimensions[col_letter].width = px_width / 7

    @staticmethod
    def _export_row_heights(ws, sheet):
        """行高をワークシートに反映"""
        row_heights = json.loads(sheet['row_heights'] or '{}')
        default_row_h = row_heights.pop('default', None)
        if default_row_h is not None:
            ws.sheet_format.defaultRowHeight = default_row_h
        for row_str, height in row_heights.items():
            ws.row_dimensions[int(row_str) + 1].height = height

    def _export_cells(self, conn, ws, sheet_name):
        """セル値・スタイルをワークシートに反映"""
        cells = conn.execute(
            "SELECT row, col, value, style FROM cells WHERE sheet_name = ?",
            (sheet_name,)
        ).fetchall()

        for cell_row in cells:
            r, c = cell_row['row'] + 1, cell_row['col'] + 1
            cell = ws.cell(row=r, column=c)
            val = cell_row['value']

            if val is not None and val != '':
                try:
                    if '.' in val:
                        cell.value = float(val)
                    else:
                        cell.value = int(val)
                except (ValueError, TypeError):
                    cell.value = val
            else:
                cell.value = val if val else None

            if cell_row['style']:
                style = json.loads(cell_row['style'])
                self._apply_openpyxl_style(cell, style)

    @staticmethod
    def _export_merges(conn, ws, sheet_name):
        """セル結合をワークシートに反映"""
        merges = conn.execute(
            "SELECT start_row, start_col, end_row, end_col FROM merges WHERE sheet_name = ?",
            (sheet_name,)
        ).fetchall()
        for m in merges:
            ws.merge_cells(
                start_row=m['start_row'] + 1, start_column=m['start_col'] + 1,
                end_row=m['end_row'] + 1, end_column=m['end_col'] + 1
            )

    @staticmethod
    def _add_anchored_image(ws, image_obj, item):
        """画像をセル位置にアンカーして配置"""
        image_obj.width = item['width']
        image_obj.height = item['height']
        col_letter = get_column_letter(item['col'] + 1)
        image_obj.anchor = f"{col_letter}{item['row'] + 1}"
        ws.add_image(image_obj)

    def _export_shapes(self, conn, ws, sheet_name):
        """図形をPNG変換してワークシートに埋め込み"""
        shapes = conn.execute(
            "SELECT * FROM shapes WHERE sheet_name = ? ORDER BY id",
            (sheet_name,)
        ).fetchall()
        for shp in shapes:
            try:
                png_bytes = self._shape_to_png_bytes(dict(shp))
                openpyxl_img = OpenpyxlImage(io.BytesIO(png_bytes))
                self._add_anchored_image(ws, openpyxl_img, shp)
            except Exception as e:
                logger.error("図形復元エラー: %s", e)

    def _export_images(self, conn, ws, sheet_name):
        """画像をワークシートに埋め込み"""
        imgs = conn.execute(
            "SELECT * FROM images WHERE sheet_name = ? ORDER BY id",
            (sheet_name,)
        ).fetchall()
        for img in imgs:
            if img['image_data']:
                try:
                    openpyxl_img = OpenpyxlImage(io.BytesIO(img['image_data']))
                    self._add_anchored_image(ws, openpyxl_img, img)
                except Exception as e:
                    logger.error("画像復元エラー: %s", e)

    def _export_charts(self, conn, ws, sheet_name):
        """グラフをワークシートに埋め込み"""
        charts = conn.execute(
            "SELECT * FROM charts WHERE sheet_name = ? ORDER BY id",
            (sheet_name,)
        ).fetchall()
        for ch in charts:
            try:
                chart_obj = self._create_openpyxl_chart(ch, ws)
                if chart_obj:
                    col_letter = get_column_letter(ch['col'] + 1)
                    cell_ref = f"{col_letter}{ch['row'] + 1}"
                    ws.add_chart(chart_obj, cell_ref)
            except Exception as e:
                logger.error("グラフ復元エラー: %s", e)

    def sqlite_to_csv(self):
        """SQLiteからCSVファイルを復元"""
        with self._connect() as conn:
            # CSVエンコーディングを取得
            row = conn.execute(
                "SELECT value FROM metadata WHERE key = 'csv_encoding'"
            ).fetchone()
            encoding = row['value'] if row else 'utf-8'

            # 最初のシートのデータを取得
            sheet = conn.execute(
                "SELECT * FROM sheets ORDER BY sheet_order LIMIT 1"
            ).fetchone()
            if not sheet:
                return

            sn = sheet['sheet_name']
            max_row = sheet['max_row']
            max_col = sheet['max_col']

            # セルデータをグリッドに復元
            grid = [[''] * max_col for _ in range(max_row)]
            cells = conn.execute(
                "SELECT row, col, value FROM cells WHERE sheet_name = ?",
                (sn,)
            ).fetchall()
            for cell in cells:
                r, c = cell['row'], cell['col']
                if r < max_row and c < max_col:
                    grid[r][c] = cell['value'] or ''

        with open(self.filepath, 'w', encoding=encoding, newline='') as f:
            writer = csv.writer(f)
            writer.writerows(grid)

    def _apply_openpyxl_style(self, cell, style):
        """スタイル辞書をopenpyxlセルに適用"""
        if 'backgroundColor' in style:
            color = style['backgroundColor'].lstrip('#')
            cell.fill = PatternFill(start_color=color, end_color=color, fill_type='solid')

        if 'border' in style:
            sides = {}
            for side_name in ['left', 'right', 'top', 'bottom']:
                if side_name in style['border']:
                    info = style['border'][side_name]
                    color = info.get('color', '#000000').lstrip('#')
                    sides[side_name] = Side(style=info.get('style', 'thin'), color=color)
            if sides:
                cell.border = Border(**sides)

        font_kwargs = {}
        if 'fontWeight' in style:
            font_kwargs['bold'] = style['fontWeight'] == 'bold'
        if 'fontStyle' in style:
            font_kwargs['italic'] = style['fontStyle'] == 'italic'
        if 'color' in style:
            color_value = style['color'].lstrip('#').upper()
            if len(color_value) == 6:
                color_value = 'FF' + color_value
            if len(color_value) == 8:
                font_kwargs['color'] = Color(rgb=color_value)
        if 'fontSize' in style:
            font_kwargs['size'] = style['fontSize']
        if font_kwargs:
            cell.font = Font(**font_kwargs)

        if 'textAlign' in style or 'verticalAlign' in style:
            cell.alignment = Alignment(
                horizontal=style.get('textAlign'),
                vertical=style.get('verticalAlign')
            )

    # ============================================================
    # 読み込み
    # ============================================================

    def read_all_sheets(self):
        """全シートデータを取得（既存JSON形式と同一）

        Returns:
            dict: {'sheets': {name: data}, 'sheetNames': [names]}
        """
        with self._connect() as conn:
            sheets = conn.execute(
                "SELECT * FROM sheets ORDER BY sheet_order"
            ).fetchall()

            sheets_data = {}
            sheet_names = []

            for sheet in sheets:
                sn = sheet['sheet_name']
                sheet_names.append(sn)
                sheets_data[sn] = self._read_sheet_data(conn, sheet)

            return {'sheets': sheets_data, 'sheetNames': sheet_names}

    def _read_sheet_data(self, conn, sheet):
        """1シート分のデータを読み込み"""
        sn = sheet['sheet_name']

        # セル
        cells = {}
        for row in conn.execute(
            "SELECT row, col, value, style FROM cells WHERE sheet_name = ?", (sn,)
        ):
            key = f"{row['row']},{row['col']}"
            style = json.loads(row['style']) if row['style'] else {}
            cells[key] = {'value': row['value'] if row['value'] is not None else '', 'style': style}

        # セル結合
        merges = [
            {
                'startRow': m['start_row'], 'startCol': m['start_col'],
                'endRow': m['end_row'], 'endCol': m['end_col']
            }
            for m in conn.execute(
                "SELECT start_row, start_col, end_row, end_col FROM merges WHERE sheet_name = ?", (sn,)
            )
        ]

        # 列幅・行高
        col_widths_raw = json.loads(sheet['col_widths'] or '{}')
        col_widths = {(k if k == 'default' else int(k)): v for k, v in col_widths_raw.items()}
        row_heights_raw = json.loads(sheet['row_heights'] or '{}')
        row_heights = {(k if k == 'default' else int(k)): v for k, v in row_heights_raw.items()}

        # オートフィルター
        auto_filter = json.loads(sheet['auto_filter']) if sheet['auto_filter'] else None

        # 画像メタデータ（画像データは遅延取得、idをユニーク識別子として使用）
        images = self._fetch_image_metadata(conn, sn)

        # 図形メタデータ
        shapes = []
        for shp in conn.execute(
            "SELECT id, shape_type, row, col, offset_x, offset_y, width, height, "
            "fill_color, stroke_color, stroke_width, opacity, rotation "
            "FROM shapes WHERE sheet_name = ? ORDER BY id", (sn,)
        ):
            shapes.append({
                'shapeIndex': shp['id'],
                'shapeType': shp['shape_type'],
                'row': shp['row'],
                'col': shp['col'],
                'offsetX': shp['offset_x'],
                'offsetY': shp['offset_y'],
                'width': shp['width'],
                'height': shp['height'],
                'fillColor': shp['fill_color'],
                'strokeColor': shp['stroke_color'],
                'strokeWidth': shp['stroke_width'],
                'opacity': shp['opacity'],
                'rotation': shp['rotation'],
            })

        # グラフメタデータ
        charts = []
        for ch in conn.execute(
            "SELECT id, chart_type, min_row, max_row, min_col, max_col, "
            "row, col, offset_x, offset_y, width, height, title, options "
            "FROM charts WHERE sheet_name = ? ORDER BY id", (sn,)
        ):
            charts.append({
                'chartIndex': ch['id'],
                'chartType': ch['chart_type'],
                'minRow': ch['min_row'],
                'maxRow': ch['max_row'],
                'minCol': ch['min_col'],
                'maxCol': ch['max_col'],
                'row': ch['row'],
                'col': ch['col'],
                'offsetX': ch['offset_x'],
                'offsetY': ch['offset_y'],
                'width': ch['width'],
                'height': ch['height'],
                'title': ch['title'],
                'options': ch['options'],
            })

        return {
            'cells': cells,
            'merges': merges,
            'colWidths': col_widths,
            'rowHeights': row_heights,
            'autoFilter': auto_filter,
            'maxRow': sheet['max_row'],
            'maxCol': sheet['max_col'],
            'images': images,
            'shapes': shapes,
            'charts': charts,
        }

    # ============================================================
    # セル更新
    # ============================================================

    def update_cells(self, sheet_name, changes):
        """セルを更新する。

        Args:
            sheet_name: シート名
            changes: [{'row': r, 'col': c, 'value': v, 'style': {...}}, ...]
        """
        with self._connect() as conn:
            for change in changes:
                r = change['row']
                c = change['col']
                value = change.get('value')
                style = change.get('style')

                existing = conn.execute(
                    "SELECT value, style FROM cells WHERE sheet_name = ? AND row = ? AND col = ?",
                    (sheet_name, r, c)
                ).fetchone()

                if existing:
                    new_value = value if value is not None else (existing['value'] or '')
                    if style:
                        old_style = json.loads(existing['style']) if existing['style'] else {}
                        old_style.update(style)
                        new_style = json.dumps(old_style)
                    else:
                        new_style = existing['style']
                    conn.execute(
                        "UPDATE cells SET value = ?, style = ? WHERE sheet_name = ? AND row = ? AND col = ?",
                        (new_value, new_style, sheet_name, r, c)
                    )
                else:
                    new_value = value if value is not None else ''
                    new_style = json.dumps(style) if style else None
                    conn.execute(
                        "INSERT INTO cells (sheet_name, row, col, value, style) VALUES (?, ?, ?, ?, ?)",
                        (sheet_name, r, c, new_value, new_style)
                    )

            # max_row / max_col 更新
            if changes:
                max_r = max(ch['row'] for ch in changes) + 1
                max_c = max(ch['col'] for ch in changes) + 1
                conn.execute(
                    "UPDATE sheets SET max_row = MAX(max_row, ?), max_col = MAX(max_col, ?) WHERE sheet_name = ?",
                    (max_r, max_c, sheet_name)
                )

            self._set_dirty(conn)

    # ============================================================
    # シート操作
    # ============================================================

    def add_sheet(self, sheet_name):
        """シートを追加"""
        with self._connect() as conn:
            existing = conn.execute(
                "SELECT sheet_name FROM sheets WHERE sheet_name = ?", (sheet_name,)
            ).fetchone()
            if existing:
                raise ValueError('同じ名前のシートが既に存在します')

            max_order = conn.execute("SELECT MAX(sheet_order) FROM sheets").fetchone()[0]
            new_order = (max_order or 0) + 1
            conn.execute(
                "INSERT INTO sheets VALUES (?, ?, 1, 1, NULL, '{}', '{}')",
                (sheet_name, new_order)
            )
            self._set_dirty(conn)

    def _cascade_delete_sheet_data(self, conn, sheet_name):
        """シートに紐づく全テーブルのデータを一括削除"""
        for table in self._SHEET_TABLES:
            conn.execute(f"DELETE FROM {table} WHERE sheet_name = ?", (sheet_name,))

    def _cascade_rename_sheet_data(self, conn, old_name, new_name):
        """シート名を全テーブルで一括変更"""
        for table in self._SHEET_TABLES:
            conn.execute(
                f"UPDATE {table} SET sheet_name = ? WHERE sheet_name = ?",
                (new_name, old_name)
            )

    def delete_sheet(self, sheet_name):
        """シートを削除"""
        with self._connect() as conn:
            count = conn.execute("SELECT COUNT(*) FROM sheets").fetchone()[0]
            if count <= 1:
                raise ValueError('シートは最低1つ必要です')

            existing = conn.execute(
                "SELECT sheet_name FROM sheets WHERE sheet_name = ?", (sheet_name,)
            ).fetchone()
            if not existing:
                raise ValueError('指定されたシートが見つかりません')

            self._cascade_delete_sheet_data(conn, sheet_name)
            self._set_dirty(conn)

    def rename_sheet(self, old_name, new_name):
        """シート名を変更"""
        with self._connect() as conn:
            existing = conn.execute(
                "SELECT sheet_name FROM sheets WHERE sheet_name = ?", (old_name,)
            ).fetchone()
            if not existing:
                raise ValueError('指定されたシートが見つかりません')

            dup = conn.execute(
                "SELECT sheet_name FROM sheets WHERE sheet_name = ?", (new_name,)
            ).fetchone()
            if dup:
                raise ValueError('同じ名前のシートが既に存在します')

            self._cascade_rename_sheet_data(conn, old_name, new_name)
            self._set_dirty(conn)

    # ============================================================
    # フィルター・列幅・行高
    # ============================================================

    def set_auto_filter(self, sheet_name, filter_ref):
        """オートフィルターを設定"""
        with self._connect() as conn:
            af = json.dumps({'ref': filter_ref}) if filter_ref else None
            conn.execute(
                "UPDATE sheets SET auto_filter = ? WHERE sheet_name = ?",
                (af, sheet_name)
            )
            self._set_dirty(conn)

    def set_column_width(self, sheet_name, col, width):
        """列幅を設定（ピクセル単位）"""
        with self._connect() as conn:
            row = conn.execute(
                "SELECT col_widths FROM sheets WHERE sheet_name = ?", (sheet_name,)
            ).fetchone()
            widths = json.loads(row['col_widths'] or '{}') if row else {}
            widths[str(col)] = width
            conn.execute(
                "UPDATE sheets SET col_widths = ? WHERE sheet_name = ?",
                (json.dumps(widths), sheet_name)
            )
            self._set_dirty(conn)

    def set_row_height(self, sheet_name, row, height):
        """行高を設定"""
        with self._connect() as conn:
            r = conn.execute(
                "SELECT row_heights FROM sheets WHERE sheet_name = ?", (sheet_name,)
            ).fetchone()
            heights = json.loads(r['row_heights'] or '{}') if r else {}
            heights[str(row)] = height
            conn.execute(
                "UPDATE sheets SET row_heights = ? WHERE sheet_name = ?",
                (json.dumps(heights), sheet_name)
            )
            self._set_dirty(conn)

    # ============================================================
    # セル結合
    # ============================================================

    def add_merges(self, sheet_name, merges):
        """セル結合を追加する。

        Args:
            sheet_name: シート名
            merges: [{'startRow': r1, 'startCol': c1, 'endRow': r2, 'endCol': c2}, ...]

        Raises:
            ValueError: 既存のマージと重複する場合、または単一セルの場合
        """
        with self._connect() as conn:
            for merge in merges:
                sr, sc = merge['startRow'], merge['startCol']
                er, ec = merge['endRow'], merge['endCol']

                if sr == er and sc == ec:
                    raise ValueError('単一セルは結合できません')

                overlapping = conn.execute(
                    "SELECT id FROM merges "
                    "WHERE sheet_name = ? "
                    "AND NOT (end_row < ? OR start_row > ? OR end_col < ? OR start_col > ?)",
                    (sheet_name, sr, er, sc, ec)
                ).fetchone()
                if overlapping:
                    raise ValueError('既存の結合セルと重複しています')

                conn.execute(
                    "INSERT INTO merges (sheet_name, start_row, start_col, end_row, end_col) "
                    "VALUES (?, ?, ?, ?, ?)",
                    (sheet_name, sr, sc, er, ec)
                )
            self._set_dirty(conn)

    def remove_merges(self, sheet_name, merges):
        """セル結合を解除する。

        Args:
            sheet_name: シート名
            merges: [{'startRow': r1, 'startCol': c1, 'endRow': r2, 'endCol': c2}, ...]
        """
        with self._connect() as conn:
            for merge in merges:
                conn.execute(
                    "DELETE FROM merges WHERE sheet_name = ? "
                    "AND start_row = ? AND start_col = ? AND end_row = ? AND end_col = ?",
                    (sheet_name, merge['startRow'], merge['startCol'],
                     merge['endRow'], merge['endCol'])
                )
            self._set_dirty(conn)

    # ============================================================
    # 画像操作
    # ============================================================

    @staticmethod
    def _fetch_image_metadata(conn, sheet_name):
        """画像メタデータ一覧を取得（内部共通処理）"""
        images = []
        for img in conn.execute(
            "SELECT id, row, col, offset_x, offset_y, width, height, "
            "image_type, stamped_at, stamped_by, rotation "
            "FROM images WHERE sheet_name = ? ORDER BY id", (sheet_name,)
        ):
            meta = {
                'imageIndex': img['id'],
                'row': img['row'],
                'col': img['col'],
                'offsetX': img['offset_x'],
                'offsetY': img['offset_y'],
                'width': img['width'],
                'height': img['height'],
                'rotation': img['rotation'] or 0,
            }
            if img['image_type']:
                meta['imageType'] = img['image_type']
            if img['stamped_at']:
                meta['stampedAt'] = img['stamped_at']
            if img['stamped_by']:
                meta['stampedBy'] = img['stamped_by']
            images.append(meta)
        return images

    def get_image_metadata(self, sheet_name):
        """シート内の画像メタデータ一覧"""
        with self._connect() as conn:
            return self._fetch_image_metadata(conn, sheet_name)

    def get_image_data(self, sheet_name, image_id):
        """画像のバイナリデータを取得（idで検索）

        Returns:
            (image_bytes, content_type) or (None, None)
        """
        with self._connect() as conn:
            img = conn.execute(
                "SELECT image_data, content_type FROM images WHERE sheet_name = ? AND id = ?",
                (sheet_name, image_id)
            ).fetchone()
            if img and img['image_data']:
                return bytes(img['image_data']), img['content_type']
            return None, None

    def save_image(self, sheet_name, image_data_dict):
        """画像を新規保存（バイナリ＋座標を紐づけて保存）

        Args:
            image_data_dict: {'row', 'col', 'offsetX', 'offsetY', 'width', 'height', 'dataUrl'}

        Returns:
            int: 画像のユニークID（SQLite auto-increment id）
        """
        target_row = image_data_dict.get('row', 0)
        target_col = image_data_dict.get('col', 0)
        width = image_data_dict.get('width', 200)
        height = image_data_dict.get('height', 200)
        data_url = image_data_dict.get('dataUrl', '')

        if not data_url.startswith('data:'):
            raise ValueError('無効な画像データです')

        header, encoded = data_url.split(',', 1)
        image_bytes = base64.b64decode(encoded)

        # MIMEタイプを判定（data:image/png;base64 形式から抽出）
        mime_match = re.search(r'image/(\w+)', header)
        content_type = f'image/{mime_match.group(1)}' if mime_match else 'image/png'

        # 画像をそのまま保存（元ピクセルサイズを維持し、表示サイズはDBのwidth/heightで管理）
        from PIL import Image as PILImage
        pil_image = PILImage.open(io.BytesIO(image_bytes))

        buf = io.BytesIO()
        fmt = 'PNG' if content_type == 'image/png' else content_type.split('/')[-1].upper()
        if fmt == 'JPEG':
            pil_image = pil_image.convert('RGB')
        pil_image.save(buf, format=fmt)
        final_bytes = buf.getvalue()

        # 押印メタデータ
        image_type = image_data_dict.get('imageType')
        stamped_at = image_data_dict.get('stampedAt')
        stamped_by = image_data_dict.get('stampedBy')

        with self._connect() as conn:
            cursor = conn.execute(
                "INSERT INTO images (sheet_name, row, col, offset_x, offset_y, width, height, "
                "content_type, image_data, image_type, stamped_at, stamped_by) "
                "VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)",
                (sheet_name, target_row, target_col,
                 int(image_data_dict.get('offsetX', 0)), int(image_data_dict.get('offsetY', 0)),
                 int(width), int(height), content_type, final_bytes,
                 image_type, stamped_at, stamped_by)
            )
            self._set_dirty(conn)
            return cursor.lastrowid

    def update_image_position(self, sheet_name, image_data_dict):
        """画像の座標・サイズ・回転のみ更新（バイナリは変更しない）

        Args:
            image_data_dict: {'imageIndex'(=id), 'row', 'col', 'offsetX', 'offsetY', 'width', 'height', 'rotation'}
        """
        image_id = image_data_dict.get('imageIndex')
        if image_id is None:
            raise ValueError('imageIndexが指定されていません')

        with self._connect() as conn:
            conn.execute(
                "UPDATE images SET row = ?, col = ?, offset_x = ?, offset_y = ?, "
                "width = ?, height = ?, rotation = ? "
                "WHERE sheet_name = ? AND id = ?",
                (
                    int(image_data_dict.get('row', 0)),
                    int(image_data_dict.get('col', 0)),
                    int(image_data_dict.get('offsetX', 0)),
                    int(image_data_dict.get('offsetY', 0)),
                    int(image_data_dict.get('width', 200)),
                    int(image_data_dict.get('height', 200)),
                    int(image_data_dict.get('rotation', 0)),
                    sheet_name, image_id
                )
            )
            self._set_dirty(conn)

    def _delete_entity(self, table_name, sheet_name, entity_id):
        """画像・図形・グラフの汎用削除"""
        with self._connect() as conn:
            conn.execute(
                f"DELETE FROM {table_name} WHERE sheet_name = ? AND id = ?",
                (sheet_name, entity_id)
            )
            self._set_dirty(conn)

    def delete_image(self, sheet_name, image_id):
        """画像を削除（idで特定）"""
        self._delete_entity('images', sheet_name, image_id)

    # ============================================================
    # 図形操作
    # ============================================================

    @staticmethod
    def _shape_params(shape_data):
        """図形データからDB用パラメータタプルを生成"""
        return (
            int(shape_data.get('row', 0)),
            int(shape_data.get('col', 0)),
            int(shape_data.get('offsetX', 0)),
            int(shape_data.get('offsetY', 0)),
            int(shape_data.get('width', 120)),
            int(shape_data.get('height', 80)),
            shape_data.get('fillColor', '#4285f4'),
            shape_data.get('strokeColor', '#333333'),
            int(shape_data.get('strokeWidth', 2)),
            float(shape_data.get('opacity', 1.0)),
            int(shape_data.get('rotation', 0)),
        )

    def save_shape(self, sheet_name, shape_data):
        """図形を新規保存

        Returns:
            int: 図形のユニークID
        """
        params = (sheet_name, shape_data.get('shapeType', 'rectangle')) + self._shape_params(shape_data)
        with self._connect() as conn:
            cursor = conn.execute(
                "INSERT INTO shapes (sheet_name, shape_type, row, col, offset_x, offset_y, "
                "width, height, fill_color, stroke_color, stroke_width, opacity, rotation) "
                "VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)",
                params
            )
            self._set_dirty(conn)
            return cursor.lastrowid

    def update_shape(self, sheet_name, shape_data):
        """図形の位置・サイズ・スタイルを更新"""
        shape_id = shape_data.get('shapeIndex')
        if shape_id is None:
            raise ValueError('shapeIndexが指定されていません')

        params = self._shape_params(shape_data) + (sheet_name, shape_id)
        with self._connect() as conn:
            conn.execute(
                "UPDATE shapes SET row = ?, col = ?, offset_x = ?, offset_y = ?, "
                "width = ?, height = ?, fill_color = ?, stroke_color = ?, "
                "stroke_width = ?, opacity = ?, rotation = ? "
                "WHERE sheet_name = ? AND id = ?",
                params
            )
            self._set_dirty(conn)

    def delete_shape(self, sheet_name, shape_id):
        """図形を削除"""
        self._delete_entity('shapes', sheet_name, shape_id)

    # ============================================================
    # グラフ操作
    # ============================================================

    @staticmethod
    def _chart_params(chart_data):
        """グラフデータからDB用パラメータタプルを生成（位置・サイズ・メタ）"""
        return (
            chart_data.get('chartType', 'bar'),
            int(chart_data.get('row', 0)),
            int(chart_data.get('col', 0)),
            int(chart_data.get('offsetX', 0)),
            int(chart_data.get('offsetY', 0)),
            int(chart_data.get('width', 480)),
            int(chart_data.get('height', 320)),
            chart_data.get('title', ''),
            chart_data.get('options', '{}'),
        )

    def save_chart(self, sheet_name, chart_data):
        """グラフを新規保存

        Returns:
            int: グラフのユニークID
        """
        cp = self._chart_params(chart_data)
        data_range = (
            int(chart_data.get('minRow', 0)),
            int(chart_data.get('maxRow', 0)),
            int(chart_data.get('minCol', 0)),
            int(chart_data.get('maxCol', 0)),
        )
        # (sheet_name, chart_type, min_row..max_col, row, col, offset_x..options)
        params = (sheet_name, cp[0]) + data_range + cp[1:]
        with self._connect() as conn:
            cursor = conn.execute(
                "INSERT INTO charts (sheet_name, chart_type, min_row, max_row, min_col, max_col, "
                "row, col, offset_x, offset_y, width, height, title, options) "
                "VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)",
                params
            )
            self._set_dirty(conn)
            return cursor.lastrowid

    def update_chart(self, sheet_name, chart_data):
        """グラフの位置・サイズ・タイプを更新"""
        chart_id = chart_data.get('chartIndex')
        if chart_id is None:
            raise ValueError('chartIndexが指定されていません')

        params = self._chart_params(chart_data) + (sheet_name, chart_id)
        with self._connect() as conn:
            conn.execute(
                "UPDATE charts SET chart_type = ?, row = ?, col = ?, offset_x = ?, offset_y = ?, "
                "width = ?, height = ?, title = ?, options = ? "
                "WHERE sheet_name = ? AND id = ?",
                params
            )
            self._set_dirty(conn)

    def delete_chart(self, sheet_name, chart_id):
        """グラフを削除"""
        self._delete_entity('charts', sheet_name, chart_id)

    def _shape_to_png_bytes(self, shape_dict):
        """図形をPNG画像バイトに変換（Excel出力用）"""
        from PIL import Image as PILImage, ImageDraw

        w = shape_dict['width']
        h = shape_dict['height']
        fill = shape_dict.get('fill_color', '#4285f4')
        stroke = shape_dict.get('stroke_color', '#333333')
        sw = shape_dict.get('stroke_width', 2)
        shape_type = shape_dict['shape_type']

        # RGBA色を解析
        def parse_color(hex_color):
            hex_color = hex_color.lstrip('#')
            if len(hex_color) == 6:
                r, g, b = int(hex_color[0:2], 16), int(hex_color[2:4], 16), int(hex_color[4:6], 16)
                return (r, g, b, 255)
            return (66, 133, 244, 255)

        fill_rgba = None if fill == 'none' else parse_color(fill)
        stroke_rgba = parse_color(stroke)

        img = PILImage.new('RGBA', (w, h), (0, 0, 0, 0))
        draw = ImageDraw.Draw(img)

        # 描画領域（ストローク幅を考慮）
        pad = sw // 2
        box = [pad, pad, w - pad - 1, h - pad - 1]

        if shape_type == 'rectangle':
            draw.rectangle(box, fill=fill_rgba, outline=stroke_rgba, width=sw)
        elif shape_type == 'rounded_rectangle':
            radius = min(w, h) // 5
            draw.rounded_rectangle(box, radius=radius, fill=fill_rgba, outline=stroke_rgba, width=sw)
        elif shape_type == 'ellipse':
            draw.ellipse(box, fill=fill_rgba, outline=stroke_rgba, width=sw)
        elif shape_type == 'triangle':
            points = [(w // 2, pad), (pad, h - pad - 1), (w - pad - 1, h - pad - 1)]
            draw.polygon(points, fill=fill_rgba, outline=stroke_rgba, width=sw)
        elif shape_type in ('line', 'arrow'):
            draw.line([(pad, h // 2), (w - pad - 1, h // 2)], fill=stroke_rgba, width=sw)
            if shape_type == 'arrow':
                # 矢印の先端
                arrow_size = max(8, sw * 3)
                ax, ay = w - pad - 1, h // 2
                points = [(ax, ay), (ax - arrow_size, ay - arrow_size // 2), (ax - arrow_size, ay + arrow_size // 2)]
                draw.polygon(points, fill=stroke_rgba)

        buf = io.BytesIO()
        img.save(buf, format='PNG')
        return buf.getvalue()

    def _create_openpyxl_chart(self, chart_row, ws):
        """SQLiteのchartレコードからopenpyxlチャートを生成"""
        chart_type = chart_row['chart_type']
        # openpyxlは1ベース
        min_row = chart_row['min_row'] + 1
        max_row = chart_row['max_row'] + 1
        min_col = chart_row['min_col'] + 1
        max_col = chart_row['max_col'] + 1

        # チャートオブジェクト生成
        chart_map = {
            'bar': BarChart,
            'line': LineChart,
            'pie': PieChart,
            'doughnut': DoughnutChart,
            'radar': RadarChart,
            'polarArea': RadarChart,
        }
        chart_cls = chart_map.get(chart_type)
        if not chart_cls:
            return None

        chart_obj = chart_cls()

        if chart_row['title']:
            chart_obj.title = chart_row['title']

        # サイズ設定（px → cm: 1px ≈ 0.0264cm）
        chart_obj.width = chart_row['width'] * 0.0264
        chart_obj.height = chart_row['height'] * 0.0264

        # データ範囲（1列目をカテゴリ、2列目以降をデータとして扱う）
        if max_col > min_col:
            cats = Reference(ws, min_col=min_col, min_row=min_row + 1, max_row=max_row)
            for col_idx in range(min_col + 1, max_col + 1):
                data = Reference(ws, min_col=col_idx, min_row=min_row, max_row=max_row)
                chart_obj.add_data(data, titles_from_data=True)
            chart_obj.set_categories(cats)
        else:
            # 1列のみの場合
            data = Reference(ws, min_col=min_col, min_row=min_row, max_row=max_row)
            chart_obj.add_data(data, titles_from_data=False)

        # polarArea → RadarChart filled スタイル
        if chart_type == 'polarArea':
            chart_obj.type = 'filled'

        return chart_obj

    # ============================================================
    # ETag / dirty
    # ============================================================

    def get_etag(self):
        """ETag用のハッシュを返す（.sqliteのmtime+sizeベース）"""
        if not os.path.exists(self.sqlite_path):
            return None
        stat = os.stat(self.sqlite_path)
        raw = f"{self.sqlite_path}:{stat.st_mtime}:{stat.st_size}"
        return hashlib.md5(raw.encode()).hexdigest()

    def mark_dirty(self):
        """ダーティフラグをセット"""
        with self._connect() as conn:
            self._set_dirty(conn)

    def _set_dirty(self, conn):
        """内部用: ダーティフラグをセット"""
        conn.execute("INSERT OR REPLACE INTO metadata VALUES ('dirty', '1')")

    def is_dirty(self):
        """ダーティフラグを確認"""
        with self._connect() as conn:
            row = conn.execute(
                "SELECT value FROM metadata WHERE key = 'dirty'"
            ).fetchone()
            return row and row['value'] == '1'

    def clear_dirty(self):
        """ダーティフラグをクリア"""
        with self._connect() as conn:
            conn.execute("INSERT OR REPLACE INTO metadata VALUES ('dirty', '0')")

    # ============================================================
    # フォルダ解析用
    # ============================================================

    def read_rows_for_analysis(self, max_rows=None):
        """フォルダ解析用: 最初のシートの行をリストのリストで返す

        Args:
            max_rows: 最大行数。Noneなら全行を返す。
        """
        with self._connect() as conn:
            sheet = conn.execute(
                "SELECT * FROM sheets ORDER BY sheet_order LIMIT 1"
            ).fetchone()
            if not sheet:
                return []

            sn = sheet['sheet_name']
            row_limit = min(sheet['max_row'], max_rows) if max_rows else sheet['max_row']
            max_col = sheet['max_col']

            grid = [[''] * max_col for _ in range(row_limit)]
            for cell in conn.execute(
                "SELECT row, col, value FROM cells "
                "WHERE sheet_name = ? AND row < ? AND col < ?",
                (sn, row_limit, max_col)
            ):
                grid[cell['row']][cell['col']] = cell['value'] or ''

            return grid

    def read_columns_for_analysis(self, col_indices, start_row=0, max_rows=None):
        """フォルダ解析用: 指定列のみを読み込む（メモリ・CPU最適化版）

        Args:
            col_indices: 読み込む列インデックスのリスト（0ベース）
            start_row: 開始行（0ベース、ヘッダー除く）
            max_rows: 最大行数。Noneなら全行を返す。

        Returns:
            list: 行データのリスト。各行は col_indices の順で値を格納。
        """
        with self._connect() as conn:
            sheet = conn.execute(
                "SELECT * FROM sheets ORDER BY sheet_order LIMIT 1"
            ).fetchone()
            if not sheet:
                return []

            sn = sheet['sheet_name']
            total_rows = sheet['max_row']
            end_row = total_rows
            if max_rows is not None:
                end_row = min(total_rows, start_row + max_rows)

            # col_indices → 内部位置マッピング
            idx_map = {c: i for i, c in enumerate(col_indices)}
            placeholders = ','.join('?' * len(col_indices))

            rows_dict = {}
            for cell in conn.execute(
                f"SELECT row, col, value FROM cells "
                f"WHERE sheet_name = ? AND row >= ? AND row < ? "
                f"AND col IN ({placeholders})",
                [sn, start_row, end_row] + list(col_indices)
            ):
                r = cell['row'] - start_row
                if r not in rows_dict:
                    rows_dict[r] = [''] * len(col_indices)
                rows_dict[r][idx_map[cell['col']]] = cell['value'] or ''

            # 連番の行リストに変換（歯抜けは空行）
            row_count = end_row - start_row
            return [rows_dict.get(i, [''] * len(col_indices))
                    for i in range(row_count)]
