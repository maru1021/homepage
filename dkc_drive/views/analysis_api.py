"""
DKCドライブ - フォルダ分析API
"""
import os
import logging
from django.views import View
from django.http import JsonResponse

from ..sqlite_converter import SpreadsheetSQLite
from .utils import (
    validate_path, get_full_path, error_response, JSONRequestMixin,
)

logger = logging.getLogger(__name__)

SPREADSHEET_EXTENSIONS = {'.xlsx', '.xls', '.csv'}


def find_table_region(rows, return_header_index=False):
    """3行以上連続でデータがある部分を表として検出

    入力rowsは全セルが文字列であることを前提とする
    （read_spreadsheet_rows経由で文字列変換済み）。

    Args:
        rows: 行データのリスト
        return_header_index: Trueの場合、ヘッダー行インデックスも返す

    Returns:
        return_header_index=False: (columns, data_rows)
        return_header_index=True:  (columns, data_rows, header_index)
    """
    empty = ([], [], 0) if return_header_index else ([], [])
    if not rows:
        return empty

    def is_data_row(row):
        """行がデータ行かどうか判定（2つ以上の空でないセルがある）"""
        if not row:
            return False
        non_empty = sum(1 for cell in row if cell and cell.strip())
        return non_empty >= 2

    # 連続したデータ行のブロックを検出
    table_start = None
    consecutive_count = 0
    best_start = None
    best_length = 0

    for i, row in enumerate(rows):
        if is_data_row(row):
            if table_start is None:
                table_start = i
            consecutive_count += 1
        else:
            # 連続が途切れた
            if consecutive_count >= 3 and consecutive_count > best_length:
                best_start = table_start
                best_length = consecutive_count
            table_start = None
            consecutive_count = 0

    # 最後まで連続している場合
    if consecutive_count >= 3 and consecutive_count > best_length:
        best_start = table_start
        best_length = consecutive_count

    if best_start is None:
        # 3行以上の連続がない場合は従来通り1行目をヘッダーとする
        if rows:
            columns = [c or '' for c in rows[0]]
            if return_header_index:
                return columns, rows[1:], 0
            return columns, rows[1:]
        return empty

    # 表の領域を抽出（コピーせずスライス参照）
    header_idx = best_start
    columns = [c or '' for c in rows[header_idx]]
    data = rows[header_idx + 1:header_idx + best_length]
    if return_header_index:
        return columns, data, header_idx
    return columns, data


def read_spreadsheet_rows(filepath, ext, max_rows=None):
    """スプレッドシートファイルから行データを読み込む。

    SQLiteがあればSQLiteから、なければCSV/Excelから直接読み込む。

    Args:
        max_rows: 最大行数。Noneなら全行を返す。

    Returns:
        list: 行データのリスト。読み込み失敗時は空リスト。
    """
    sqlite_path = filepath + '.sqlite'
    if os.path.exists(sqlite_path):
        try:
            db = SpreadsheetSQLite(filepath)
            return db.read_rows_for_analysis(max_rows=max_rows)
        except Exception:
            logger.debug("SQLite読み込みフォールバック: %s", filepath)

    if ext == '.csv':
        import csv
        with open(filepath, 'r', encoding='utf-8-sig', errors='replace') as f:
            reader = csv.reader(f)
            if max_rows:
                return [row for _, row in zip(range(max_rows), reader)]
            return list(reader)

    # Excel
    try:
        from openpyxl import load_workbook
        wb = load_workbook(filepath, read_only=True, data_only=True)
        ws = wb.active
        rows = [[str(c) if c is not None else '' for c in row]
                for row in ws.iter_rows(values_only=True, max_row=max_rows)]
        wb.close()
        return rows
    except Exception as e:
        logger.warning("Excel読み込みエラー: %s", e)
        return []


def _list_spreadsheet_files(target_dir):
    """フォルダ内のスプレッドシートファイルを収集（.sqliteのみも対象、重複排除済み）

    Returns:
        list of (display_name, filepath, ext)
    """
    candidates = []
    seen = set()
    for name in sorted(os.listdir(target_dir)):
        item_path = os.path.join(target_dir, name)
        if os.path.isdir(item_path):
            continue
        if name.endswith('.sqlite'):
            display_name = name[:-7]
            ext = os.path.splitext(display_name)[1].lower()
            if ext in SPREADSHEET_EXTENSIONS and display_name not in seen:
                seen.add(display_name)
                candidates.append((display_name, os.path.join(target_dir, display_name), ext))
        else:
            ext = os.path.splitext(name)[1].lower()
            if ext in SPREADSHEET_EXTENSIONS and name not in seen:
                seen.add(name)
                candidates.append((name, item_path, ext))
    return candidates


# 列名判別に必要な最大行数（FolderAnalysisAPI / FolderAnalysisDataAPI 共用）
HEADER_SCAN_ROWS = 10


class FolderAnalysisAPI(View):
    """フォルダ内のExcel/CSVから列情報を取得するAPI"""

    def get(self, request):
        folder_path = request.GET.get('folder', '')

        if not validate_path(folder_path):
            return error_response('無効なパスです')

        target_dir = get_full_path(folder_path)
        if not os.path.exists(target_dir) or not os.path.isdir(target_dir):
            return error_response('フォルダが見つかりません', 404)

        spreadsheet_files = _list_spreadsheet_files(target_dir)
        data_files, columns, all_data = self._load_headers(
            spreadsheet_files)

        return JsonResponse({
            'status': 'success',
            'folder': folder_path,
            'files': data_files,
            'columns': sorted(list(columns)),
            'allData': all_data
        })

    # ---------- ヘッダーのみ読み込み（軽量版） ----------

    def _load_headers(self, spreadsheet_files):
        data_files = []
        columns = set()
        all_data = {}

        for display_name, item_path, ext in spreadsheet_files:
            try:
                file_columns, row_count = self._read_file_columns(
                    item_path, ext)
                if file_columns:
                    data_files.append({
                        'name': display_name,
                        'columns': file_columns,
                        'rowCount': row_count,
                    })
                    columns.update(file_columns)
                    all_data[display_name] = {'columns': file_columns}
            except Exception as e:
                logger.warning(
                    "ファイル読み込みエラー: %s - %s", display_name, e)
                continue

        return data_files, columns, all_data

    def _read_file_columns(self, filepath, ext):
        """先頭行のみ読み込んで列名とデータ行数を返す"""
        rows = read_spreadsheet_rows(
            filepath, ext, max_rows=HEADER_SCAN_ROWS)
        file_columns, _ = find_table_region(rows)

        # データ行数はSQLiteのmax_rowから取得（高速）
        row_count = 0
        sqlite_path = filepath + '.sqlite'
        if os.path.exists(sqlite_path):
            try:
                db = SpreadsheetSQLite(filepath)
                with db._connect() as conn:
                    sheet = conn.execute(
                        "SELECT max_row FROM sheets "
                        "ORDER BY sheet_order LIMIT 1"
                    ).fetchone()
                    if sheet:
                        row_count = sheet['max_row'] - 1  # ヘッダー行を除く
            except Exception:
                pass

        return file_columns, row_count


class FolderAnalysisDataAPI(JSONRequestMixin, View):
    """フォルダ解析用のデータ取得API（列インデックス指定・最適化版）

    フロントエンドから指定列インデックスのみを受け取り、
    SQLiteから該当列だけをSELECTすることでCPU・メモリ負荷を削減する。
    """

    def post(self, request):
        data = self.parse_json(request)
        folder_path = data.get('folder', '')
        col_indices = data.get('col_indices', [])
        start_row = max(int(data.get('start_row', 1)), 1)
        end_row = data.get('end_row')  # None=最終行まで

        if not validate_path(folder_path):
            return error_response('無効なパスです')

        if not col_indices or not isinstance(col_indices, list):
            return error_response('列が指定されていません')

        # 整数に変換・バリデーション
        try:
            col_indices = [int(c) for c in col_indices]
        except (ValueError, TypeError):
            return error_response('列インデックスが不正です')

        target_dir = get_full_path(folder_path)
        if not os.path.exists(target_dir) or not os.path.isdir(target_dir):
            return error_response('フォルダが見つかりません', 404)

        spreadsheet_files = _list_spreadsheet_files(target_dir)
        all_data = {}

        for display_name, item_path, ext in spreadsheet_files:
            try:
                result = self._read_file_columns(
                    item_path, ext, col_indices, start_row, end_row)
                if result:
                    all_data[display_name] = result
            except Exception as e:
                logger.warning("データ抽出エラー: %s - %s", display_name, e)

        return JsonResponse({
            'status': 'success',
            'allData': all_data,
            'col_indices': col_indices,
        })

    def _read_file_columns(self, filepath, ext, col_indices, start_row, end_row):
        """指定列のデータを最適化読み込み"""
        header_rows = read_spreadsheet_rows(
            filepath, ext, max_rows=HEADER_SCAN_ROWS)
        columns, _, header_offset = find_table_region(
            header_rows, return_header_index=True)
        if not columns:
            return None

        sqlite_path = filepath + '.sqlite'
        if os.path.exists(sqlite_path):
            db = SpreadsheetSQLite(filepath)
            # header_offset行がヘッダー、データは header_offset+1 から
            data_start = header_offset + start_row  # 0ベース
            max_rows = (int(end_row) - start_row + 1) if end_row else None
            data = db.read_columns_for_analysis(
                col_indices, start_row=data_start, max_rows=max_rows)
        else:
            # SQLiteがない場合のフォールバック
            all_rows = read_spreadsheet_rows(filepath, ext)
            _, table_data = find_table_region(all_rows)
            s = start_row - 1
            e = int(end_row) if end_row else len(table_data)
            data = [
                [(row[ci] if ci < len(row) else '') for ci in col_indices]
                for row in table_data[s:e]
            ]

        return {'columns': columns, 'data': data}
